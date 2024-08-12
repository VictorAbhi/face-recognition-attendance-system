import os
from datetime import datetime
import time
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


class AttendanceLogger:
    def __init__(self, attendance_dir='Attendance', col_names=['NAME', 'TIME']):
        self.attendance_dir = attendance_dir
        self.col_names = col_names

        # Create the attendance directory if it doesn't exist
        if not os.path.exists(self.attendance_dir):
            os.makedirs(self.attendance_dir)

    def log_attendance(self, name):
        ts = time.time()
        date = datetime.fromtimestamp(ts).strftime("%d-%m-%y")
        timestamp = datetime.fromtimestamp(ts).strftime("%H:%M:%S")

        attendance = [name, timestamp]
        attendance_file = os.path.join(self.attendance_dir, f'Attendance_{date}.xlsx')

        if os.path.isfile(attendance_file):
            workbook = load_workbook(attendance_file)
            sheet = workbook.active
        else:
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(self.col_names)  # Write the header only if file doesn't exist

        # Append the new attendance record
        sheet.append(attendance)

        # Auto-size columns
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column].width = adjusted_width

        # Save the workbook
        workbook.save(attendance_file)